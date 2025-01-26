"""_summary_

    Raises:
        ValueError: _description_

    Returns:
        _type_: _description_
    """
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import time
import openpyxl


def browse_master_file():
    """Opens a file dialog to select the master spreadsheet."""
    filepath = filedialog.askopenfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
    )
    if filepath:
        master_file_entry.delete(0, tk.END)
        master_file_entry.insert(0, filepath)


def browse_folder():
    """Opens a file dialog to select the folder."""
    folder_path = filedialog.askdirectory()
    if folder_path:
        folder_entry.delete(0, tk.END)
        folder_entry.insert(0, folder_path)


def update_spreadsheet(master_file_path, new_file_path):
    """Updates the master spreadsheet with data from a new file."""
    try:
        master_wb = openpyxl.load_workbook(master_file_path)
        master_sheet = master_wb.active

        new_wb = openpyxl.load_workbook(new_file_path)
        new_sheet = new_wb.active

        # Get headers from the master sheet
        headers = [cell.value for cell in master_sheet[1]]
        if "CRN" not in headers:
            raise ValueError("The 'CRN' column is missing from the master spreadsheet.")

        crn_index = headers.index("CRN")

        # Create a set of CRNs from the master sheet
        master_crns = {
            row[crn_index]
            for row in master_sheet.iter_rows(min_row=2, values_only=True)
            if row[crn_index] is not None
        }

        # Update or add rows based on the new file
        for new_row in new_sheet.iter_rows(min_row=2, values_only=True):
            if new_row[crn_index] in master_crns:
                # Update existing rows
                for row in master_sheet.iter_rows(min_row=2):
                    if row[crn_index].value == new_row[crn_index]:
                        for col, value in enumerate(new_row, start=1):
                            row[col - 1].value = value
                        break
            else:
                # Append new rows
                master_sheet.append(new_row)

        master_wb.save(master_file_path)
        return True

    except FileNotFoundError as e:
        messagebox.showerror("Error", f"File not found: {e}")
    except openpyxl.utils.exceptions.InvalidFileException as e:
        messagebox.showerror("Error", f"Invalid Excel file: {e}")
    except ValueError as e:
        messagebox.showerror("Error", str(e))
    except PermissionError as e:
        messagebox.showerror("Error", f"File permission error: {e}")
    except OSError as e:
        messagebox.showerror("Error", f"OS error: {e}")
    return False


def update_progress(current_file, total_files):
    """Updates the progress bar and status label."""
    progress_bar["value"] = (current_file / total_files) * 100
    status_label.config(text=f"Processing file {current_file} of {total_files}")
    root.update_idletasks()


def start_update():
    """Starts the spreadsheet update process in a separate thread."""
    master_file_path = master_file_entry.get()
    folder_path = folder_entry.get()

    if not master_file_path or not folder_path:
        messagebox.showerror(
            "Error", "Please select both the master file and the folder."
        )
        return

    def update_thread():
        try:
            files = [
                f
                for f in os.listdir(folder_path)
                if f.endswith(".xlsx") and f != os.path.basename(master_file_path)
            ]
            total_files = len(files)

            for i, filename in enumerate(files):
                new_file_path = os.path.join(folder_path, filename)
                if update_spreadsheet(master_file_path, new_file_path):
                    update_progress(i + 1, total_files)
                    time.sleep(1)
                else:
                    break

            status_label.config(text="Update complete!")
        except FileNotFoundError as e:
            messagebox.showerror("Error", f"File not found: {e}")
        except PermissionError as e:
            messagebox.showerror("Error", f"File permission error: {e}")
        except OSError as e:
            messagebox.showerror("Error", f"OS error: {e}")

    threading.Thread(target=update_thread).start()


# --- GUI Setup ---
root = tk.Tk()
root.title("Spreadsheet Updater")

# Master spreadsheet selection
master_file_label = tk.Label(root, text="Master Spreadsheet:")
master_file_label.grid(row=0, column=0, padx=5, pady=5)
master_file_entry = tk.Entry(root, width=50)
master_file_entry.grid(row=0, column=1, padx=5, pady=5)
master_file_button = tk.Button(root, text="Browse", command=browse_master_file)
master_file_button.grid(row=0, column=2, padx=5, pady=5)

# Folder selection
folder_label = tk.Label(root, text="Folder:")
folder_label.grid(row=1, column=0, padx=5, pady=5)
folder_entry = tk.Entry(root, width=50)
folder_entry.grid(row=1, column=1, padx=5, pady=5)
folder_button = tk.Button(root, text="Browse", command=browse_folder)
folder_button.grid(row=1, column=2, padx=5, pady=5)

# Update button
update_button = tk.Button(root, text="Update Now", command=start_update)
update_button.grid(row=2, column=1, padx=5, pady=10)

# Progress bar
progress_bar = ttk.Progressbar(
    root, orient="horizontal", length=300, mode="determinate"
)
progress_bar.grid(row=3, column=0, columnspan=3, padx=5, pady=5)

# Status label
status_label = tk.Label(root, text="")
status_label.grid(row=4, column=0, columnspan=3, padx=5, pady=5)

root.mainloop()
